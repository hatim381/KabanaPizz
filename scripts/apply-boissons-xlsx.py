import json
import re
import ssl
import urllib.request
from pathlib import Path
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DRINKS_DIR = ROOT / "public" / "drinks"
MENU_JS = ROOT / "src" / "data" / "menu.js"
XLSX = ROOT / "boissons.xlsx"

NAME_TO_ID = {
    "Coca-Cola 33 cl": "coca-cola-33",
    "Coca-Cola 1,25 L": "coca-cola-125",
    "Coca-Cola Cherry 33 cl": "coca-cherry-33",
    "Coca Zero 33 cl": "coca-zero-33",
    "Coca-Cola Zero 1,25 L": "coca-zero-125",
    "Fanta 1,25 L": "fanta-125",
    "Oasis Tropical 33 cl": "oasis-tropical-33",
    "Oasis Tropical 2 L": "oasis-tropical-2l",
    "Oasis Cassis 33 cl": "oasis-cassis-33",
    "Orangina 33 cl": "orangina-33",
    "Perrier 33 cl": "perrier-33",
    "Canada Dry 33 cl": "canada-dry-33",
    "Tropico 33 cl": "tropico-33",
    "Ice Tea 33 cl": "ice-tea-33",
    "Lipton Ice Tea 1,25 L": "lipton-ice-tea-125",
    "Red Bull 25 cl": "redbull-25",
    "Eau 50 cl": "eau-50",
    "Eau plate 1,5 L": "eau-plate-150",
}

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) KabanaPizzSite/1.0"
IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def fetch(url: str) -> tuple[bytes, str]:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept-Language": "fr-FR,fr;q=0.9"})
    for ctx in (ssl.create_default_context(), ssl._create_unverified_context()):
        try:
            with urllib.request.urlopen(req, timeout=30, context=ctx) as res:
                data = res.read()
                ctype = res.headers.get("Content-Type", "")
                final_url = res.geturl()
                return data, ctype if ctype else final_url
        except urllib.error.HTTPError:
            raise
        except Exception:
            if ctx.verify_mode == ssl.CERT_NONE:
                raise
            continue
    raise RuntimeError(f"échec téléchargement {url}")


def is_direct_image(url: str, content_type: str, data: bytes | None = None) -> bool:
    if "svg" in content_type.lower():
        return False
    if data and data[:5].startswith(b"<?xml") or (data and data[:4] == b"<svg"):
        return False
    if content_type.startswith("image/"):
        return True
    path = urlparse(url).path.lower()
    return any(path.endswith(ext) for ext in IMAGE_EXT) and not path.endswith(".svg")


def candidate_images(html: str, base_url: str) -> list[str]:
    patterns = [
        r'<meta[^>]+property=["\']og:image(?::secure_url)?["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image(?::secure_url)?["\']',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']twitter:image["\']',
        r'"image"\s*:\s*"(https?://[^"]+\.(?:jpg|jpeg|png|webp))"',
        r'itemprop=["\']image["\'][^>]+src=["\']([^"\']+)["\']',
        r'<img[^>]+src=["\']([^"\']+\.(?:jpg|jpeg|png|webp))["\']',
        r'data-src=["\']([^"\']+\.(?:jpg|jpeg|png|webp))["\']',
    ]
    seen = set()
    out = []
    for pat in patterns:
        for m in re.finditer(pat, html, re.I):
            url = urljoin(base_url, m.group(1).replace("\\/", "/"))
            low = url.lower()
            if url in seen or ".svg" in low or "logo" in low:
                continue
            seen.add(url)
            out.append(url)
    return out


def resolve_image_url(url: str) -> tuple[str, bytes, str]:
    data, meta = fetch(url)
    if is_direct_image(url, meta, data):
        return url, data, meta
    html = data.decode("utf-8", errors="ignore")
    for img_url in candidate_images(html, url):
        try:
            img_data, img_meta = fetch(img_url)
            if is_direct_image(img_url, img_meta, img_data):
                return img_url, img_data, img_meta
        except Exception:
            continue
    raise ValueError(f"aucune image trouvée sur {url}")


def pick_ext(url: str, content_type: str) -> str:
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    if "gif" in content_type:
        return ".gif"
    path = urlparse(url).path.lower()
    for ext in IMAGE_EXT:
        if path.endswith(ext):
            return ".jpg" if ext == ".jpeg" else ext
    return ".jpg"


def update_menu_js(updates: dict[str, str]) -> None:
    text = MENU_JS.read_text(encoding="utf-8")
    for drink_id, image_path in updates.items():
        pattern = rf'(id: "{re.escape(drink_id)}", name: .+?, price: [^,]+, image: )"[^"]*"'
        new_text, n = re.subn(pattern, rf'\1"{image_path}"', text, count=1)
        if n != 1:
            raise ValueError(f"impossible de mettre à jour {drink_id} dans menu.js")
        text = new_text
    MENU_JS.write_text(text, encoding="utf-8")


def main():
    wb = load_workbook(XLSX)
    sheet = wb.active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    new_col = None
    for i, h in enumerate(headers, start=1):
        if h and "nouvelle" in str(h).lower():
            new_col = i
            break
    if not new_col:
        raise SystemExit("Colonne 'Nouvelle image url' introuvable")

    name_col = headers.index("Nom de la boisson") + 1
    results = []
    menu_updates = {}

    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value
        url = sheet.cell(row, new_col).value
        if not name or not url:
            continue
        drink_id = NAME_TO_ID.get(str(name).strip())
        if not drink_id:
            print(f"SKIP {name}: id inconnu")
            continue
        try:
            img_url, data, ctype = resolve_image_url(str(url).strip())
            ext = pick_ext(img_url, ctype)
            out = DRINKS_DIR / f"{drink_id}{ext}"
            out.write_bytes(data)
            rel = f"/drinks/{drink_id}{ext}"
            menu_updates[drink_id] = rel
            results.append({"id": drink_id, "name": name, "source": url, "resolved": img_url, "file": str(out)})
            print(f"OK  {drink_id} <- {img_url}")
        except Exception as e:
            print(f"ECHEC {drink_id}: {e}")
            results.append({"id": drink_id, "name": name, "source": url, "error": str(e)})

    if menu_updates:
        update_menu_js(menu_updates)
        print(f"\nmenu.js mis à jour ({len(menu_updates)} boissons)")

    report = ROOT / "scripts" / "boissons-import-report.json"
    report.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Rapport: {report}")


if __name__ == "__main__":
    main()